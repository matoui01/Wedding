#!/usr/bin/env python3
"""Turn the couple's working spreadsheet into a seeded Guests tab.

    python3 tools/invites/sheet/migrate-xlsx.py <Wedding_invitee.xlsx> [out.csv]

The source list is one row per invitation already, which is the shape the
engine wants. What changes here:

  * the single `Category` column splits into Category + Subcategory + Side, so
    the list can be sliced by band ("all Family") or by circle ("everyone from
    Geneva") without either getting in the other's way;
  * `PlusOne` splits into a yes/no and a name, with placeholders ("femme",
    "copine") dropped — an invitation should not address someone as "femme";
  * `Language` is guessed from the circle, since it is far quicker to correct a
    wrong guess than to fill ninety blanks;
  * the pivot tables, the `Flaky` scratch column and nameless rows are dropped;
  * households listed only in the loose right-hand column, which no total on
    the sheet counts, are brought in as Hold so they surface for a decision
    instead of quietly vanishing.

`Email` is deliberately left empty — it is the one thing only the couple can
fill, and the engine refuses to draft a row without it.
"""
import csv
import pathlib
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

GUEST_HEADER = [
    'Household', 'Invitee', 'Plus-one', 'Plus-one name', 'Kids?', 'Kids est.',
    'Category', 'Subcategory', 'Side', 'Priority', 'Send?', 'Wave',
    'Email', 'Language', 'Greeting', 'Personal note',
    'Seats', 'Token', 'Invite link', 'Reply by', 'Invite status', 'Invite sent',
    'Reminder sent', 'RSVP', 'Adults', 'Children', 'Coming', 'Diet', 'Shuttle', 'RSVP at',
]

# old Category -> (Category, Subcategory)
BANDS = {
    'ATOUI':             ('Family', 'Atoui'),
    'BAUDOIN':           ('Family', 'Baudoin'),
    'Close Family':      ('Close Family', ''),
    'Family Friends':    ('Family Friends', ''),
    'Distant Family':    ('Family', 'Distant family'),
    'Friends GVA':       ('Friends', 'Geneva'),
    'Friends GVA - ITA': ('Friends', 'Geneva · Italian'),
    'Friends Italy':     ('Friends', 'Italy'),
    'Friends Prepa':     ('Friends', 'Prépa'),
    'Friends School':    ('Friends', 'Engineering school'),
}

# "Sophie" is a person; "femme" is a role. Only the first belongs in an invite.
PLACEHOLDERS = {
    'femme', 'copine', 'copain', 'mari', 'marito', 'moglie', 'compagne',
    'compagnon', 'conjoint', 'conjointe', 'son conjoint', 'sa femme',
    'son mari', 'girlfriend', 'boyfriend', 'wife', 'husband', '+1', '1',
}

LANG_BY_SUB = {
    'Geneva': 'fr', 'Geneva · Italian': 'it', 'Italy': 'it',
    'Prépa': 'fr', 'Engineering school': 'fr', 'Atoui': 'fr', 'Baudoin': 'fr',
}
LANG_BY_SIDE = {'Ilaria': 'it', 'Maxime': 'fr', 'Common': 'en'}

# rows where the source contradicts itself: one seat, but a name in PlusOne
REVIEW = []

# headings from the pivot tables that share the right-hand columns
PIVOT_LABELS = {
    'row labels', 'column labels', 'grand total', 'common', 'ilaria', 'maxime',
    '(blank)', 'sum of #',
}


def S(v):
    return '' if v is None else str(v).strip()


def guess_lang(sub, side):
    """A friend of both of them has no obvious mother tongue, so "Common" wins
    over the city and gets English. Otherwise the circle is the better signal,
    with the side as fallback. A wrong guess costs one dropdown click."""
    if side == 'Common':
        return 'en'
    return LANG_BY_SUB.get(sub) or LANG_BY_SIDE.get(side, 'en')


def split_plus_one(raw, seats):
    """-> (yes/no, name). Numbers and role-words carry no name.

    `#` is the authority on how many seats a row is for. The PlusOne column is
    sometimes a surname instead of a partner ("Martina Mariotti | Pelle"), so
    where the two disagree the count wins and the text is kept for review —
    better a name to check than an invented guest."""
    raw = S(raw)
    numeric = bool(re.fullmatch(r'\d+(\.\d+)?', raw))
    if seats <= 1:
        return ('no', '' if numeric else raw)
    if not raw or numeric or raw.lower() in PLACEHOLDERS:
        return ('yes', '')
    return ('yes', raw)


def household(invitee, plus, plus_name):
    """Already-joined names ("Tarek et sa femme") stay as they are."""
    if re.search(r'\s(et|e|and|&)\s', invitee, re.I):
        return invitee
    return f'{invitee} & {plus_name}' if (plus == 'yes' and plus_name) else invitee


def row(invitee, plus_raw, seats, cat, sub, side, priority, send):
    plus, plus_name = split_plus_one(plus_raw, seats)
    if plus == 'no' and plus_name:
        REVIEW.append((invitee, plus_name))
    out = {h: '' for h in GUEST_HEADER}
    out.update({
        'Household': household(invitee, plus, plus_name), 'Invitee': invitee,
        'Plus-one': plus, 'Plus-one name': plus_name,
        'Category': cat, 'Subcategory': sub, 'Side': side,
        'Priority': priority, 'Send?': send,
        'Language': guess_lang(sub, side),
    })
    return out


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = pathlib.Path(sys.argv[1])
    out = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else \
        pathlib.Path(__file__).with_name('guests-seed.csv')

    ws = openpyxl.load_workbook(src, data_only=True)['Sheet1']
    rows, seen, unknown = [], set(), set()

    for r in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        _, n, priority, old_cat, side, invitee, plus, _flaky = r
        invitee, old_cat, side = S(invitee), S(old_cat), S(side)
        if not invitee or not old_cat:
            continue                       # spacer rows and stray totals
        if old_cat not in BANDS:
            unknown.add(old_cat)
            continue
        cat, sub = BANDS[old_cat]
        seats = int(n) if isinstance(n, (int, float)) else 1
        rows.append(row(invitee, plus, seats, cat, sub, side, S(priority), 'Send'))
        seen.add(invitee.lower())

    # The loose column-L list: the source these family blocks were typed from.
    # Anything there that never made it into the table is real, uncounted, and
    # needs a decision — so it comes in on Hold rather than being assumed.
    extra = 0
    for i in range(2, ws.max_row + 1):
        name, count = S(ws.cell(i, 12).value), ws.cell(i, 14).value
        if not name or not isinstance(count, (int, float)):
            continue
        # the pivot tables sit in these same columns; their labels and totals
        # are not people, so anything without a letter in it is not a name
        if not re.search(r'[^\W\d_]', name) or name.lower() in PIVOT_LABELS:
            continue
        if name.lower() in seen:
            continue
        seats = int(count)
        rows.append(row(name, '1' if seats >= 2 else '', seats,
                        '', '', 'Maxime', '', 'Hold'))
        seen.add(name.lower())
        extra += 1

    with out.open('w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=GUEST_HEADER)
        w.writeheader()
        w.writerows(rows)

    planned = sum(1 for x in rows if x['Send?'] == 'Send')
    seats = sum(1 + (x['Plus-one'] == 'yes') for x in rows)
    print(f'{out}: {len(rows)} households, {seats} seats')
    print(f'  {planned} ready to send · {extra} held for classification')
    if unknown:
        print(f'  skipped unmapped categories: {sorted(unknown)}')
    if REVIEW:
        print(f'  {len(REVIEW)} row(s) counted as one seat but carrying a name in '
              f'PlusOne — kept in Plus-one name, please check:')
        for inv, nm in REVIEW:
            print(f'    {inv} + {nm!r}')


if __name__ == '__main__':
    main()
